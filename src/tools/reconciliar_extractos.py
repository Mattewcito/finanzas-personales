"""
reconciliar_extractos.py
=========================
Extrae movimientos de los extractos PDF (cuenta de ahorros + tarjetas de
crédito Mastercard/AMEX) y los compara contra finanzas_personales.xlsx
para detectar cuáles NO están registrados todavía.

Uso:
    python reconciliar_extractos.py            -> solo genera el REPORTE
                                                   (no toca el Excel)
    python reconciliar_extractos.py --aplicar  -> aplica los cambios:
                                                   agrega las filas nuevas
                                                   a "movimientos" y una
                                                   fila a
                                                   "historial_actualizaciones"

Decisiones de modelado (ver resumen impreso al final):
  - El archivo "..._Credito_AMEX_4112.pdf" es un duplicado exacto de
    "..._AMEX_4112.pdf" (mismo número de cuenta, mismo período) -> se
    ignora para no contar dos veces.
  - En los extractos de tarjeta solo se toman los "Nuevos movimientos"
    de cada período; la sección "Movimientos antes de..." son compras a
    cuotas de períodos anteriores repetidas para seguimiento, no gasto
    nuevo -> se ignoran.
  - Los intereses de ahorros (ABONO/AJUSTE INTERESES AHORROS) se
    consolidan en UNA fila por extracto (categoria "intereses") en vez
    de una fila por día -> evita cientos de filas de centavos.
  - Los intereses corrientes de tarjeta de crédito también se
    consolidan en una fila por extracto tarjeta, categoria "intereses",
    marcada como deuda de tarjeta.
  - Un pago a tarjeta de crédito puede aparecer DOS veces (una vez como
    salida en el extracto de ahorros "TRANSFERENCIA CTA SUC VIRTUAL" y
    otra vez como entrada "ABONO SUCURSAL VIRTUAL" en el extracto de la
    tarjeta) -> si coinciden fecha+monto entre mis propios candidatos,
    se conserva solo la versión de la tarjeta (categoria
    pago_tarjeta_credito) y se descarta la de ahorros.
  - La comparación contra el Excel existente usa (fecha, monto
    redondeado) como llave -- NO la descripción, porque el texto del
    banco y el texto de las alertas de Gmail describen la misma
    transacción con palabras distintas.
"""

import re
import sys
import json
import datetime
import argparse
from pathlib import Path
from collections import defaultdict

import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill

for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

DOWNLOADS = Path(r"C:\Users\User\Downloads")
XLSX_PATH = Path(__file__).resolve().parents[2] / "data" / "finanzas_personales.xlsx"  # src/tools/ -> raíz del proyecto -> data/
REPORT_PATH = Path(__file__).resolve().parent / "reconciliacion_extractos.log"

SAVINGS_FILES = [
    "Cuenta-de-ahorros-2025-12-31_2026-03-31.pdf",
    "Cuenta-de-ahorros-2025-03-31_2026-06-30.pdf",
]
CARD_FILES = [
    ("Extracto_1144681774_202608_TARJETA_MASTERCARD_2011.pdf", "2011"),
    ("Extracto_1136656136_202607_TARJETA_AMEX_4112.pdf", "4112"),
    ("Extracto_1113453234_202606_TARJETA_AMEX_4112.pdf", "4112"),
    # Extracto_1113453234_202606_TARJETA_Credito_AMEX_4112.pdf -- OMITIDO: duplicado exacto del anterior
]

MESES_ABREV = {"ene":1,"feb":2,"mar":3,"abr":4,"may":5,"jun":6,"jul":7,"ago":8,"sep":9,"oct":10,"nov":11,"dic":12}


def pdf_text(path):
    pages = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            pages.append(page.extract_text() or "")
    return pages


# ---------------------------------------------------------------------------
# CATEGORIZACIÓN (heurística por palabras clave en la descripción)
# ---------------------------------------------------------------------------
CATEGORIA_KEYWORDS = [
    ("salario", ["nomi", "nomina", "n\u00f3mina"]),
    ("transporte", ["uber", "didi", "cabify", "picap"]),
    ("comida", ["rappi", "tostao", "frisby", "dunkin", "oxxo", "la migueria", "hamburgues", "wasabi", "crepes y w", "atmos"]),
    ("restaurantes", ["sr wok", "pato pekin", "il forno", "ponto brasileiro"]),
    ("supermercado", ["tienda d1", "exito", "ara ", "tiendas ara", "jumbo", "d1 bodega"]),
    ("compras", ["mercadopago", "mercado pago", "temu", "dollarcity", "miniso", "homecenter", "alkomprar", "decathlon"]),
    ("ropa", ["lenesens", "topara", "croydon", "seven seven"]),
    ("celular", ["comcel", "claro movil", "claro m\u00f3vil", "movistar"]),
    ("internet", ["claro hogar"]),
    ("servicios", ["epm ", "empresas publicas"]),
    ("suscripciones", ["apple.com", "spotify", "netfli", "anthropic", "claude sub"]),
    ("educacion", ["corp univ iberoameri", "ibero", "corp univ"]),
    ("comida", ["la migueri"]),
    ("salud", ["mediPiel".lower(), "drogueria", "droguer\u00eda"]),
    ("entretenimiento", ["cine colombia", "escape xiv"]),
    ("mascotas", ["my tienda pets", "peluditos", "astro mascota"]),
    ("retiro", ["retiro cajero", "atm "]),
    ("transporte", ["recarga de tarjeta civica", "recarga tarjeta civica", "pago qr transporte"]),
    ("tecnologia", ["mundo digital"]),
    ("hogar", ["444 atmos"]),
]


def inferir_categoria(desc_upper, default="otros"):
    d = desc_upper.lower()
    for cat, keywords in CATEGORIA_KEYWORDS:
        for kw in keywords:
            if kw in d:
                return cat
    if "transf" in d or "transferencia" in d:
        return "transferencias"
    if "pago qr" in d or d.startswith("pago qr"):
        return "otros"
    if "consignacion" in d or "consignaci\u00f3n" in d:
        return "transferencias"
    return default


def to_float_us(s):
    """'1,009,412.17' -> 1009412.17 (coma miles, punto decimal)."""
    return float(s.replace(",", ""))


def to_float_latam(s):
    """'1.897.320,00' -> 1897320.00 (punto miles, coma decimal)."""
    return float(s.replace(".", "").replace(",", "."))


# ---------------------------------------------------------------------------
# PARSER: CUENTA DE AHORROS
# ---------------------------------------------------------------------------
LINE_RE_SAVINGS = re.compile(r"^(\d{1,2}/\d{1,2})\s+(.+?)\s+(-?[\d,]+\.\d{2})\s+([\d,]+\.\d{2})$")


def parse_savings_statement(path):
    """Devuelve (movimientos, intereses_total, periodo_desde, periodo_hasta)."""
    pages = pdf_text(path)
    full_text = "\n".join(pages)

    m = re.search(r"DESDE:\s*(\d{4})/(\d{2})/(\d{2})\s*HASTA:\s*(\d{4})/(\d{2})/(\d{2})", full_text)
    y_desde, m_desde, d_desde = int(m.group(1)), int(m.group(2)), int(m.group(3))
    y_hasta, m_hasta, d_hasta = int(m.group(4)), int(m.group(5)), int(m.group(6))
    periodo_desde = f"{y_desde:04d}-{m_desde:02d}-{d_desde:02d}"
    periodo_hasta = f"{y_hasta:04d}-{m_hasta:02d}-{d_hasta:02d}"

    movimientos = []
    intereses_total = 0.0

    for line in full_text.splitlines():
        line = line.strip()
        mo = LINE_RE_SAVINGS.match(line)
        if not mo:
            continue
        fecha_dm, desc, valor_s, _saldo_s = mo.groups()
        day, month = fecha_dm.split("/")
        day, month = int(day), int(month)
        # El estado de cuenta va de periodo_desde a periodo_hasta; todas las fechas
        # de movimiento caen en ese rango (usamos el año del extremo cuyo mes coincide).
        year = y_hasta if month <= m_hasta and (month >= m_desde or y_desde != y_hasta) else y_desde
        # Ambos extractos usados no cruzan un cambio de año dentro de sus movimientos,
        # así que year = y_hasta siempre que el mes esté en [1, m_hasta]; si el extracto
        # empezara en diciembre y el movimiento fuera de diciembre, usaría y_desde.
        if month == 12 and m_desde == 12:
            year = y_desde
        else:
            year = y_hasta
        fecha = f"{year:04d}-{month:02d}-{day:02d}"

        desc_up = desc.strip()
        valor = to_float_us(valor_s)

        if "INTERESES AHORROS" in desc_up.upper() or "AJUSTE INTERES" in desc_up.upper():
            intereses_total += valor
            continue

        movimientos.append({
            "fecha": fecha,
            "descripcion_original": desc_up,
            "valor": valor,   # + = abono (ingreso), - = cargo (gasto)
        })

    return movimientos, intereses_total, periodo_desde, periodo_hasta


# ---------------------------------------------------------------------------
# PARSER: TARJETAS DE CRÉDITO (Mastercard / AMEX)
# ---------------------------------------------------------------------------
LINE_RE_CARD = re.compile(
    r"(?:^|\s)(\d{2}/\d{2}/\d{4})\s+([A-Z0-9ÁÉÍÓÚÑ][A-Z0-9ÁÉÍÓÚÑ\.\*/ ]*?)\s+\$\s*(-?[\d\.]+,\d{2})"
)


def parse_card_statement(path, ultimos4):
    pages = pdf_text(path)
    movimientos = []
    intereses_total_por_moneda = defaultdict(float)
    periodo_desde = periodo_hasta = None

    for page_text in pages:
        # Detecta la moneda de la página (cada tarjeta reporta 2 páginas: PESOS y DOLARES)
        if "ESTADO DE CUENTA EN: DOLARES" in page_text or "MMMooonnneeedddaaa::: DDDOOOLLLAAARRREEESSS" in page_text:
            moneda = "USD"
        elif "ESTADO DE CUENTA EN: PESOS" in page_text or "MMMooonnneeedddaaa::: PPPEEESSSOOOSSS" in page_text:
            moneda = "COP"
        else:
            moneda = None

        if "DDDeeetttaaalllllleeesss dddeeelll mmmooovvviiimmmiiieeennntttooo" not in page_text and "Detalles del movimiento" not in page_text:
            continue  # página de resumen/carátula, no de detalle de movimientos

        # Solo la sección de "Nuevos movimientos" -- cortamos antes de "Movimientos antes de"
        corte = re.split(r"[Mm]{3}[o0]{3}[v]{3}[i]{3}[m]{3}[i]{3}[e]{3}[n]{3}[t]{3}[o]{3}[s]{3}\s+[a]{3}[n]{3}[t]{3}[e]{3}[s]{3}", page_text)
        # fallback simple si el patrón triplicado no aplica (texto normal)
        if len(corte) == 1:
            corte = re.split(r"Movimientos antes de", page_text)
        seccion_nueva = corte[0]

        pm = re.search(r"entre\s+(\d{2}/\d{2}/\d{4})\s+hasta\s+(\d{2}/\d{2}/\d{4})", seccion_nueva, re.IGNORECASE)
        if pm:
            d0, d1 = pm.groups()
            periodo_desde = min(periodo_desde, d0) if periodo_desde else d0
            periodo_hasta = max(periodo_hasta, d1) if periodo_hasta else d1

        for line in seccion_nueva.splitlines():
            line = line.strip()
            if not line or moneda is None:
                continue
            mo = LINE_RE_CARD.search(line)
            if not mo:
                continue
            fecha_s, desc, valor_s = mo.groups()
            dd, mm, yyyy = fecha_s.split("/")
            fecha = f"{yyyy}-{mm}-{dd}"
            desc_up = desc.strip().rstrip(".")
            valor = to_float_latam(valor_s)

            if "INTERESES CORRIENTES" in desc_up.upper():
                intereses_total_por_moneda[(fecha, moneda)] += valor
                continue

            movimientos.append({
                "fecha": fecha,
                "descripcion_original": desc_up,
                "valor": valor,       # + = cargo/compra (deuda), - = abono/pago a la tarjeta
                "moneda": moneda,
                "ultimos4": ultimos4,
            })

    return movimientos, intereses_total_por_moneda, periodo_desde, periodo_hasta


# ---------------------------------------------------------------------------
# NORMALIZACIÓN a esquema del Excel
# ---------------------------------------------------------------------------
def normalizar_savings(mov, entidad="Bancolombia"):
    valor = mov["valor"]
    desc = mov["descripcion_original"]
    tipo = "ingreso" if valor > 0 else "gasto"
    categoria = inferir_categoria(desc)
    if tipo == "ingreso" and categoria == "otros":
        categoria = "transferencias" if ("transf" in desc.lower() or "consignacion" in desc.lower() or "consignaci\u00f3n" in desc.lower()) else categoria
    if desc.upper().startswith("PAGO DE NOMI"):
        categoria = "salario"
    return {
        "fecha": mov["fecha"],
        "tipo": tipo,
        "categoria": categoria,
        "moneda": "COP",
        "monto": round(abs(valor), 2),
        "descripcion": desc.title() if desc.isupper() else desc,
        "entidad": entidad,
        "_fuente": "ahorros",
        "_valor_signo": valor,
    }


def normalizar_card(mov, marca):
    valor = mov["valor"]
    desc = mov["descripcion_original"]
    ultimos4 = mov["ultimos4"]
    moneda = mov["moneda"]

    if "AVANCE SUCURSAL" in desc.upper():
        tipo = "gasto"  # el enriquecedor del dashboard lo reclasifica a ingreso/deuda automáticamente
        categoria = "credito"
        desc_fmt = f"Avance T.Cred *{ultimos4} a cta *5360"
    elif "ABONO SUCURSAL" in desc.upper() or "APLICACION SALDO" in desc.upper() or "TRASLADO SALDO" in desc.upper():
        tipo = "gasto"
        categoria = "pago_tarjeta_credito"
        desc_fmt = f"Pago tarjeta {marca} *{ultimos4}"
    elif "COMISION AVANCE" in desc.upper():
        tipo = "gasto"
        categoria = "credito"
        desc_fmt = f"Comision avance T.Cred *{ultimos4}"
    else:
        tipo = "gasto"
        categoria = inferir_categoria(desc)
        desc_fmt = f"Compra en {desc.title()} con T.Cred *{ultimos4}"

    return {
        "fecha": mov["fecha"],
        "tipo": tipo,
        "categoria": categoria,
        "moneda": moneda,
        "monto": round(abs(valor), 2),
        "descripcion": desc_fmt,
        "entidad": "Bancolombia",
        "_fuente": f"tarjeta_{marca}_{ultimos4}",
        "_valor_signo": valor,
    }


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--aplicar", action="store_true", help="Escribe los cambios en el Excel (por defecto solo reporta).")
    args = parser.parse_args()

    candidatos = []
    resumen_fuentes = []

    # --- Cuentas de ahorro ---
    for fname in SAVINGS_FILES:
        path = DOWNLOADS / fname
        movs, intereses, p_desde, p_hasta = parse_savings_statement(path)
        for m in movs:
            candidatos.append(normalizar_savings(m))
        if abs(intereses) > 0.001:
            mes_repr = p_hasta[:7]
            candidatos.append({
                "fecha": p_hasta, "tipo": "ingreso", "categoria": "intereses",
                "moneda": "COP", "monto": round(intereses, 2),
                "descripcion": f"Intereses ahorro acumulados {p_desde} a {p_hasta}",
                "entidad": "Bancolombia", "_fuente": "ahorros_intereses", "_valor_signo": intereses,
            })
        resumen_fuentes.append((fname, len(movs), p_desde, p_hasta))

    # --- Tarjetas de crédito ---
    for fname, ultimos4 in CARD_FILES:
        path = DOWNLOADS / fname
        marca = "Mastercard" if ultimos4 == "2011" else "Amex"
        movs, intereses_por_moneda, p_desde, p_hasta = parse_card_statement(path, ultimos4)
        for m in movs:
            candidatos.append(normalizar_card(m, marca))
        for (fecha_interes, moneda), val in intereses_por_moneda.items():
            if abs(val) < 0.001:
                continue
            candidatos.append({
                "fecha": fecha_interes, "tipo": "gasto", "categoria": "intereses",
                "moneda": moneda, "monto": round(val, 2),
                "descripcion": f"Interes corriente T.Cred *{ultimos4}",
                "entidad": "Bancolombia", "_fuente": f"tarjeta_{marca}_intereses", "_valor_signo": val,
            })
        resumen_fuentes.append((fname, len(movs), p_desde, p_hasta))

    # --- Dedup interno: pago a tarjeta puede aparecer 2 veces (ahorros + tarjeta) ---
    pagos_tarjeta = [c for c in candidatos if c["categoria"] == "pago_tarjeta_credito"]
    transferencias_ahorros = [c for c in candidatos if c["_fuente"] == "ahorros" and c["categoria"] == "transferencias" and c["tipo"] == "gasto"]
    duplicados_cruzados = set()
    for pago in pagos_tarjeta:
        for i, tr in enumerate(transferencias_ahorros):
            if tr["fecha"] == pago["fecha"] and abs(tr["monto"] - pago["monto"]) < 1:
                duplicados_cruzados.add(id(tr))
                break
    candidatos = [c for c in candidatos if id(c) not in duplicados_cruzados]

    # --- Dedup interno: mismo (fecha, monto, moneda) repetido entre mis propios candidatos ---
    vistos = set()
    candidatos_unicos = []
    for c in candidatos:
        k = (c["fecha"], c["moneda"], round(c["monto"], 0))
        if k in vistos:
            continue
        vistos.add(k)
        candidatos_unicos.append(c)
    candidatos = candidatos_unicos

    # --- Cargar Excel existente y construir índice de dedup ---
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["movimientos"]
    existentes_keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        fecha, tipo, categoria, moneda, monto, descripcion, entidad = row[:7]
        if isinstance(fecha, (datetime.datetime, datetime.date)):
            fecha = fecha.strftime("%Y-%m-%d")
        else:
            fecha = str(fecha)
        existentes_keys.add((fecha, str(moneda), round(float(monto or 0), 0)))

    nuevos = [c for c in candidatos if (c["fecha"], c["moneda"], round(c["monto"], 0)) not in existentes_keys]
    nuevos.sort(key=lambda c: c["fecha"])

    # --- Reporte ---
    print("=" * 70)
    print("RESUMEN DE EXTRACCIÓN")
    print("=" * 70)
    for fname, n, p0, p1 in resumen_fuentes:
        print(f"  {fname}: {n} movimientos ({p0} a {p1})")
    print()
    print(f"Total candidatos extraídos (tras dedup interno): {len(candidatos)}")
    print(f"Ya existen en el Excel: {len(candidatos) - len(nuevos)}")
    print(f"NUEVOS a agregar: {len(nuevos)}")
    print()
    print("Por categoría:")
    por_cat = defaultdict(lambda: [0, 0.0])
    for n in nuevos:
        por_cat[n["categoria"]][0] += 1
        por_cat[n["categoria"]][1] += n["monto"] if n["moneda"] == "COP" else 0
    for cat, (cnt, total) in sorted(por_cat.items(), key=lambda x: -x[1][1]):
        print(f"  {cat:22s} {cnt:3d} mov.  ${total:>14,.0f} COP")
    print()
    print("Rango de fechas de los nuevos:", nuevos[0]["fecha"] if nuevos else "-", "a", nuevos[-1]["fecha"] if nuevos else "-")
    print()
    print("Primeros 20 nuevos:")
    for n in nuevos[:20]:
        print(f"  {n['fecha']} {n['tipo']:8s} {n['categoria']:20s} {n['moneda']} {n['monto']:>12,.2f}  {n['descripcion']}")
    if len(nuevos) > 20:
        print(f"  ... y {len(nuevos)-20} más")

    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        json.dump([{k: v for k, v in n.items() if not k.startswith("_")} for n in nuevos], f, ensure_ascii=False, indent=2)
    print()
    print(f"Detalle completo de los {len(nuevos)} nuevos guardado en: {REPORT_PATH}")

    if not args.aplicar:
        print()
        print(">>> Modo REPORTE (no se modificó el Excel). Corré con --aplicar para escribir los cambios. <<<")
        return

    # --- Aplicar cambios al Excel ---
    import shutil
    backup_path = XLSX_PATH.with_name(XLSX_PATH.stem + f"_backup_{datetime.date.today().isoformat()}.xlsx")
    shutil.copy2(XLSX_PATH, backup_path)
    print(f"\nBackup creado en: {backup_path}")

    wb2 = openpyxl.load_workbook(XLSX_PATH)  # sin data_only, para no perder fórmulas/estilos
    ws2 = wb2["movimientos"]
    wh2 = wb2["historial_actualizaciones"]

    COLOR_INGRESO = "E8F5E9"
    COLOR_GASTO = "FFEBEE"
    next_row = ws2.max_row + 1
    for n in nuevos:
        r = next_row
        ws2.cell(r, 1, n["fecha"])
        ws2.cell(r, 2, n["tipo"])
        ws2.cell(r, 3, n["categoria"])
        ws2.cell(r, 4, n["moneda"])
        ws2.cell(r, 5, n["monto"])
        ws2.cell(r, 5).number_format = "#,##0.00"
        ws2.cell(r, 6, n["descripcion"])
        ws2.cell(r, 7, n["entidad"])
        bg = COLOR_INGRESO if n["tipo"] == "ingreso" else COLOR_GASTO
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for c in ws2[r]:
            c.fill = fill
        next_row += 1

    fechas_nuevas = [n["fecha"] for n in nuevos]
    hist_row = wh2.max_row + 1
    wh2.cell(hist_row, 1, datetime.date.today().isoformat())
    wh2.cell(hist_row, 2, min(fechas_nuevas) if fechas_nuevas else "")
    wh2.cell(hist_row, 3, max(fechas_nuevas) if fechas_nuevas else "")
    wh2.cell(hist_row, 4, len(nuevos))

    wb2.save(XLSX_PATH)
    print(f"\n✅ {len(nuevos)} movimientos agregados a 'movimientos'.")
    print(f"✅ Fila agregada a 'historial_actualizaciones'.")


if __name__ == "__main__":
    main()
