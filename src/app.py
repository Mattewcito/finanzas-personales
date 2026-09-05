"""
app.py
========
Servidor local (Flask, open-source/gratis) que sirve el dashboard, el
formulario de registro manual, la carga de extractos, y ahora también el
login multiusuario.

Uso:
    py app.py
Abre solo en el navegador: http://127.0.0.1:5001

Variables de entorno opcionales (para correr dentro de Docker):
  HOST                -> por defecto 127.0.0.1 (solo esta PC). En Docker se
                          usa 0.0.0.0 para que el contenedor sea alcanzable
                          desde afuera vía el puerto mapeado.
  PORT                -> por defecto 5001.
  RUNNING_IN_DOCKER=1  -> evita intentar abrir un navegador (no existe
                          dentro del contenedor).

Usuarios: cada movimiento pertenece a un usuario_id. El rol 'admin' puede
"cambiar de perfil" (ver los datos de cualquier cuenta); el rol 'usuario'
siempre ve solo los suyos. Ver crear_usuario.py para dar de alta cuentas.

Rutas:
  /login, /logout      -> autenticación
  /                     -> Dashboard (embebido, del perfil que se esté viendo)
  /vista/dashboard      -> el archivo pre-generado de ESE perfil
  /registrar            -> formulario de alta manual de un movimiento
  /cargar-extractos     -> formulario para subir un extracto (Excel/PDF)
  /cambiar-vista        -> (solo admin) cambia a qué perfil se está viendo
"""

import os
import sys
import secrets
import datetime
import threading
import webbrowser
from functools import wraps
from pathlib import Path

from flask import Flask, render_template, request, jsonify, send_from_directory, send_file, session, redirect, url_for

for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

import db_finanzas as db
import actualizar_dashboard

sys.path.insert(0, str(Path(__file__).resolve().parent / "tools"))
import reconciliar_extractos as rex  # parsers de PDF ya construidos y probados

UPLOADS_DIR = db.DATA_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

# Etiqueta que se muestra en el menú ("Mathewcito · Local/Online"). Ya NO
# se puede inferir de "¿estoy en Docker?" porque dev y online corren los
# dos en contenedores -- ahora cada docker-compose (override, específico
# de cada carpeta) fija su propio MODO_LABEL. "Local" es el default para
# cuando se corre directo con "py app.py" sin Docker.
MODO = os.environ.get("MODO_LABEL", "Local")

app = Flask(__name__)

# Clave de sesión: si no existe se genera una vez y se guarda en data/
# (fuera de git). Sin esto, Flask no puede firmar las cookies de sesión
# de forma segura.
_SECRET_KEY_PATH = db.DATA_DIR / "secret_key.txt"
if not _SECRET_KEY_PATH.exists():
    _SECRET_KEY_PATH.write_text(secrets.token_hex(32), encoding="utf-8")
app.secret_key = _SECRET_KEY_PATH.read_text(encoding="utf-8").strip()


# ----------------------------- Autenticación -----------------------------

def login_required(f):
    @wraps(f)
    def decorado(*args, **kwargs):
        if not session.get("usuario_id"):
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return decorado


def viendo_id() -> int:
    """El usuario_id cuyos datos hay que mostrar/afectar ahora mismo.
    Para rol 'usuario' es siempre el suyo. Para 'admin' puede ser el de
    cualquiera (ver /cambiar-vista) -- por defecto, el suyo propio."""
    return session.get("viendo_id", session.get("usuario_id"))


@app.context_processor
def inyectar_globales():
    usuarios_disponibles = []
    if session.get("rol") == "admin":
        conn = db.conectar()
        try:
            usuarios_disponibles = db.listar_usuarios(conn)
        finally:
            conn.close()
    return {
        "modo": MODO,
        "usuario_nombre": session.get("nombre"),
        "usuario_rol": session.get("rol"),
        "viendo_id": viendo_id(),
        "usuarios_disponibles": usuarios_disponibles,
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html", error=None)

    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")

    conn = db.conectar()
    try:
        db.crear_esquema(conn)
        cuenta = db.verificar_login(conn, username, password)
    finally:
        conn.close()

    if not cuenta:
        return render_template("login.html", error="Usuario o contraseña incorrectos.")

    session.clear()
    session["usuario_id"] = cuenta["id"]
    session["rol"] = cuenta["rol"]
    session["nombre"] = cuenta["nombre_mostrado"]
    session["viendo_id"] = cuenta["id"]  # por defecto, cada quien ve lo suyo al entrar
    return redirect(request.args.get("next") or url_for("home"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/cambiar-vista", methods=["POST"])
@login_required
def cambiar_vista():
    if session.get("rol") != "admin":
        return jsonify(ok=False, error="Solo el admin puede cambiar de perfil."), 403
    nuevo_id = request.form.get("usuario_id", type=int)
    conn = db.conectar()
    try:
        existe = db.obtener_usuario(conn, nuevo_id)
    finally:
        conn.close()
    if not existe:
        return jsonify(ok=False, error="Ese usuario no existe."), 400
    session["viendo_id"] = nuevo_id
    return redirect(url_for("home"))


# ----------------------------- Rutas de la app -----------------------------

@app.route("/health")
def health():
    """Sin login a propósito: la usa el pipeline de despliegue para
    confirmar que el contenedor arrancó bien antes de darlo por bueno."""
    return jsonify(ok=True), 200


@app.route("/")
@login_required
def home():
    return render_template("dashboard.html", activo="dashboard")


@app.route("/vista/dashboard")
@login_required
def vista_dashboard():
    return send_from_directory(db.DATA_DIR, f"dashboard_{viendo_id()}.html")


@app.route("/registrar")
@login_required
def registrar():
    conn = db.conectar()
    try:
        db.crear_esquema(conn)
        categorias = db.obtener_categorias(conn, usuario_id=viendo_id())
        entidades = db.obtener_entidades(conn, usuario_id=viendo_id())
    finally:
        conn.close()
    return render_template("registrar.html", activo="registrar", categorias=categorias, entidades=entidades)


@app.route("/api/registrar-movimiento", methods=["POST"])
@login_required
def api_registrar_movimiento():
    fecha = request.form.get("fecha", "").strip()
    tipo = request.form.get("tipo", "gasto").strip()
    monto_raw = request.form.get("monto", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    categoria = request.form.get("categoria", "").strip() or "otros"
    moneda = request.form.get("moneda", "COP").strip()
    entidad = request.form.get("entidad", "").strip() or "Manual"

    if not fecha or not descripcion:
        return jsonify(ok=False, error="Falta fecha o descripción."), 400
    try:
        monto = float(monto_raw)
    except ValueError:
        return jsonify(ok=False, error="El monto no es un número válido."), 400
    if monto <= 0:
        return jsonify(ok=False, error="El monto tiene que ser mayor a 0."), 400

    movimiento = {
        "fecha": fecha, "tipo": tipo, "categoria": categoria,
        "moneda": moneda, "monto": monto, "descripcion": descripcion, "entidad": entidad,
    }

    conn = db.conectar()
    try:
        db.crear_esquema(conn)
        stats = db.insertar_movimientos(conn, [movimiento], origen="manual", usuario_id=viendo_id())
    finally:
        conn.close()

    if stats["nuevos"] > 0:
        actualizar_dashboard.main()

    return jsonify(ok=True, **stats)


@app.route("/cargar-extractos")
@login_required
def cargar_extractos():
    return render_template("cargar_extractos.html", activo="cargar")


@app.route("/crear-usuario")
@login_required
def crear_usuario_page():
    if session.get("rol") != "admin":
        return redirect(url_for("home"))
    conn = db.conectar()
    try:
        usuarios = db.listar_usuarios(conn)
    finally:
        conn.close()
    return render_template("crear_usuario.html", activo="crear_usuario", usuarios=usuarios)


@app.route("/api/crear-usuario", methods=["POST"])
@login_required
def api_crear_usuario():
    if session.get("rol") != "admin":
        return jsonify(ok=False, error="Solo un administrador puede crear usuarios."), 403

    username = request.form.get("username", "").strip()
    nombre = request.form.get("nombre", "").strip()
    password = request.form.get("password", "")
    rol = request.form.get("rol", "usuario").strip()

    if not username or not nombre or not password:
        return jsonify(ok=False, error="Faltan campos."), 400
    if rol not in ("admin", "usuario"):
        return jsonify(ok=False, error="Rol inválido."), 400
    if len(password) < 4:
        return jsonify(ok=False, error="La contraseña es demasiado corta."), 400

    conn = db.conectar()
    try:
        db.crear_esquema(conn)
        nuevo_id = db.crear_usuario(conn, username, password, rol, nombre)
    finally:
        conn.close()

    actualizar_dashboard.main()  # genera el dashboard (vacío) del usuario nuevo de una vez

    return jsonify(ok=True, id=nuevo_id)


@app.route("/mi-perfil")
@login_required
def mi_perfil():
    """Edita SIEMPRE la cuenta con la que se inició sesión (no la que se
    esté 'viendo' si sos admin) -- para no confundir "mi perfil" con la
    cuenta ajena que un admin puede estar mirando en ese momento."""
    conn = db.conectar()
    try:
        cuenta = db.obtener_usuario(conn, session["usuario_id"])
    finally:
        conn.close()
    return render_template("mi_perfil.html", activo="mi_perfil", cuenta=cuenta)


@app.route("/api/actualizar-perfil", methods=["POST"])
@login_required
def api_actualizar_perfil():
    username = request.form.get("username", "").strip()
    nombre = request.form.get("nombre", "").strip()
    password = request.form.get("password", "").strip()

    if not username or not nombre:
        return jsonify(ok=False, error="Faltan campos."), 400
    if password and len(password) < 4:
        return jsonify(ok=False, error="La contraseña es demasiado corta."), 400

    conn = db.conectar()
    try:
        db.actualizar_usuario(conn, session["usuario_id"], username=username,
                               nombre_mostrado=nombre, password=password or None)
    finally:
        conn.close()

    session["nombre"] = nombre  # para que el menú lo refleje ya, sin tener que volver a loguearse

    return jsonify(ok=True)


@app.route("/editar-usuario/<int:usuario_id>")
@login_required
def editar_usuario_page(usuario_id):
    if session.get("rol") != "admin":
        return redirect(url_for("home"))
    conn = db.conectar()
    try:
        cuenta = db.obtener_usuario(conn, usuario_id)
    finally:
        conn.close()
    if not cuenta:
        return redirect(url_for("crear_usuario_page"))
    return render_template("editar_usuario.html", activo="crear_usuario", cuenta=cuenta)


@app.route("/api/editar-usuario/<int:usuario_id>", methods=["POST"])
@login_required
def api_editar_usuario(usuario_id):
    if session.get("rol") != "admin":
        return jsonify(ok=False, error="Solo un administrador puede editar otras cuentas."), 403

    username = request.form.get("username", "").strip()
    nombre = request.form.get("nombre", "").strip()
    password = request.form.get("password", "").strip()
    rol = request.form.get("rol", "").strip()

    if not username or not nombre or rol not in ("admin", "usuario"):
        return jsonify(ok=False, error="Faltan campos o rol inválido."), 400
    if password and len(password) < 4:
        return jsonify(ok=False, error="La contraseña es demasiado corta."), 400

    conn = db.conectar()
    try:
        db.actualizar_usuario(conn, usuario_id, username=username, nombre_mostrado=nombre,
                               password=password or None, rol=rol)
    finally:
        conn.close()

    if usuario_id == session.get("usuario_id"):  # el admin se editó a sí mismo
        session["nombre"] = nombre
        session["rol"] = rol

    return jsonify(ok=True)


@app.route("/plantilla-excel")
@login_required
def plantilla_excel():
    """Excel vacío (con un ejemplo) en el formato exacto que espera la
    opción "Excel" de Cargar extractos -- para que nadie tenga que
    adivinar las columnas."""
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "movimientos"
    ws.append(list(db.CAMPOS_ESPERADOS))
    ws.append(["2026-01-15", "gasto", "comida", "COP", 25000, "Ejemplo: Mercado", "Bancolombia"])
    ws.append(["2026-01-15", "ingreso", "salario", "COP", 2000000, "Ejemplo: Nomina", "Bancolombia"])

    for col in ws.columns:
        letra = col[0].column_letter
        ancho = max(len(str(c.value)) for c in col) + 2
        ws.column_dimensions[letra].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="plantilla_movimientos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/cargar-extracto", methods=["POST"])
@login_required
def api_cargar_extracto():
    archivo = request.files.get("archivo")
    tipo = request.form.get("tipo", "")
    if not archivo or not archivo.filename:
        return jsonify(ok=False, error="No se recibió ningún archivo."), 400

    destino = UPLOADS_DIR / archivo.filename
    archivo.save(destino)

    try:
        if tipo == "excel":
            movimientos = _leer_excel_generico(destino)
        elif tipo == "pdf_ahorros":
            crudos = rex.parse_savings_statement(destino)
            movimientos = [rex.normalizar_savings(m) for m in crudos]
        elif tipo == "pdf_tarjeta":
            marca = (request.form.get("marca") or "Credito").strip()
            ultimos4 = (request.form.get("ultimos4") or "").strip()
            if not ultimos4:
                return jsonify(ok=False, error="Falta indicar los últimos 4 dígitos de la tarjeta."), 400
            crudos = rex.parse_card_statement(destino, ultimos4)
            movimientos = [rex.normalizar_card(m, marca) for m in crudos]
        else:
            return jsonify(ok=False, error=f"Tipo de archivo desconocido: {tipo!r}"), 400
    except Exception as e:
        return jsonify(ok=False, error=f"No pude leer el archivo: {e}"), 400

    if not movimientos:
        return jsonify(ok=False, error="No encontré movimientos en ese archivo."), 400

    conn = db.conectar()
    try:
        db.crear_esquema(conn)
        stats = db.insertar_movimientos(conn, movimientos, origen=f"upload_{tipo}", usuario_id=viendo_id())
    finally:
        conn.close()

    actualizar_dashboard.main()  # regenera los dashboards de todos los usuarios

    return jsonify(ok=True, **stats)


def _leer_excel_generico(path) -> list[dict]:
    """Lee un .xlsx con las mismas columnas que la tabla movimientos
    (fecha, tipo, categoria, moneda, monto, descripcion, entidad), en la
    primera hoja del archivo."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h else h for h in row]
            continue
        if row[0] is None:
            continue
        d = dict(zip(headers, row))

        fecha = d.get("fecha")
        if isinstance(fecha, (datetime.datetime, datetime.date)):
            d["fecha"] = fecha.strftime("%Y-%m-%d")
        elif fecha is not None:
            d["fecha"] = str(fecha)

        monto = d.get("monto")
        d["monto"] = float(monto) if monto is not None else 0.0

        for campo in db.CAMPOS_ESPERADOS:
            d.setdefault(campo, "")

        rows.append(d)
    return rows


if __name__ == "__main__":
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", 5001))

    if not os.environ.get("RUNNING_IN_DOCKER"):
        threading.Timer(1.0, lambda: webbrowser.open(f"http://127.0.0.1:{port}")).start()

    app.run(host=host, port=port, debug=False)
