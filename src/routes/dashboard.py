"""
routes/dashboard.py
=====================
Blueprint del dashboard: verlo, registrar un movimiento a mano, y cargar
extractos (Excel/PDF). Ver auth.py para el patrón de Blueprints elegido.
"""
import datetime
import io
import sys
from pathlib import Path

from flask import Blueprint, render_template, request, jsonify, send_from_directory, send_file

import db_finanzas as db
import perfil_financiero
from auth import login_required, viendo_id

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "tools"))
import reconciliar_extractos as rex  # parsers de PDF ya construidos y probados

UPLOADS_DIR = db.DATA_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

# La plantilla del dashboard ya NO se "hornea" con datos por usuario --
# es un único archivo estático que pide sus datos por AJAX a
# /api/dashboard-data al cargar (ver dashboard/dashboard_finanzas.html,
# función iniciarDashboard()). Ver el commit que introdujo esto para el
# porqué del cambio: el modelo viejo (un dashboard_<id>.html generado
# por usuario, escrito en disco) es lo que causaba que dev/producción
# se pisaran entre sí al compartir esa carpeta, y necesitaba un paso de
# "regenerar" manual cada vez que cambiaban los datos.
DASHBOARD_DIR = db.PROJECT_ROOT / "dashboard"

dashboard_bp = Blueprint("dashboard", __name__)


@dashboard_bp.route("/")
@login_required
def home():
    return render_template("dashboard.html", activo="dashboard")


@dashboard_bp.route("/vista/dashboard")
@login_required
def vista_dashboard():
    return send_from_directory(DASHBOARD_DIR, "dashboard_finanzas.html")


@dashboard_bp.route("/api/dashboard-data")
@login_required
def api_dashboard_data():
    """Todo lo que el dashboard necesita para el perfil que se esté
    viendo ahora mismo (respeta viendo_id(), igual que cualquier otra
    ruta): movimientos, ledger de deuda, y el perfil financiero por
    hábitos -- calculado al vuelo, no desde un archivo pre-generado."""
    with db.conexion() as conn:
        movimientos = db.obtener_movimientos(conn, usuario_id=viendo_id())
        ledger_deuda = db.obtener_ledger_deuda(conn, usuario_id=viendo_id())
    perfil = perfil_financiero.generar_perfil(movimientos, ledger_deuda)
    return jsonify(
        movimientos=movimientos,
        ledger_deuda=ledger_deuda,
        perfil=perfil,
        generated_at=datetime.datetime.now().isoformat(timespec="seconds"),
    )


@dashboard_bp.route("/registrar")
@login_required
def registrar():
    with db.conexion() as conn:
        db.crear_esquema(conn)
        categorias = db.obtener_categorias(conn, usuario_id=viendo_id())
        entidades = db.obtener_entidades(conn, usuario_id=viendo_id())
    return render_template("registrar.html", activo="registrar", categorias=categorias, entidades=entidades)


@dashboard_bp.route("/api/registrar-movimiento", methods=["POST"])
@login_required
def api_registrar_movimiento():
    fecha = request.form.get("fecha", "").strip()
    tipo = request.form.get("tipo", "gasto").strip()
    monto_raw = request.form.get("monto", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    categoria = request.form.get("categoria", "").strip() or "otros"
    moneda = request.form.get("moneda", "COP").strip()
    entidad = request.form.get("entidad", "").strip() or "Manual"

    if not fecha or not descripcion:
        return jsonify(ok=False, error="Falta fecha o descripción."), 400
    try:
        monto = float(monto_raw)
    except ValueError:
        return jsonify(ok=False, error="El monto no es un número válido."), 400
    if monto <= 0:
        return jsonify(ok=False, error="El monto tiene que ser mayor a 0."), 400

    movimiento = {
        "fecha": fecha, "tipo": tipo, "categoria": categoria,
        "moneda": moneda, "monto": monto, "descripcion": descripcion, "entidad": entidad,
    }

    with db.conexion() as conn:
        db.crear_esquema(conn)
        stats = db.insertar_movimientos(conn, [movimiento], origen="manual", usuario_id=viendo_id())

    # El dashboard ya no se regenera a mano -- /api/dashboard-data lee la
    # BD en cada carga de página, así que este movimiento ya está
    # disponible ni bien el navegador vuelva a pedirlo.
    return jsonify(ok=True, **stats)


@dashboard_bp.route("/cargar-extractos")
@login_required
def cargar_extractos():
    return render_template("cargar_extractos.html", activo="cargar")


@dashboard_bp.route("/plantilla-excel")
@login_required
def plantilla_excel():
    """Excel vacío (con un ejemplo) en el formato exacto que espera la
    opción "Excel" de Cargar extractos -- para que nadie tenga que
    adivinar las columnas."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "movimientos"
    ws.append(list(db.CAMPOS_ESPERADOS))
    ws.append(["2026-01-15", "gasto", "comida", "COP", 25000, "Ejemplo: Mercado", "Bancolombia"])
    ws.append(["2026-01-15", "ingreso", "salario", "COP", 2000000, "Ejemplo: Nomina", "Bancolombia"])

    for col in ws.columns:
        letra = col[0].column_letter
        ancho = max(len(str(c.value)) for c in col) + 2
        ws.column_dimensions[letra].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="plantilla_movimientos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@dashboard_bp.route("/api/cargar-extracto", methods=["POST"])
@login_required
def api_cargar_extracto():
    archivo = request.files.get("archivo")
    tipo = request.form.get("tipo", "")
    if not archivo or not archivo.filename:
        return jsonify(ok=False, error="No se recibió ningún archivo."), 400

    destino = UPLOADS_DIR / archivo.filename
    archivo.save(destino)

    try:
        if tipo == "excel":
            movimientos = _leer_excel_generico(destino)
        elif tipo == "pdf_ahorros":
            crudos = rex.parse_savings_statement(destino)
            movimientos = [rex.normalizar_savings(m) for m in crudos]
        elif tipo == "pdf_tarjeta":
            marca = (request.form.get("marca") or "Credito").strip()
            ultimos4 = (request.form.get("ultimos4") or "").strip()
            if not ultimos4:
                return jsonify(ok=False, error="Falta indicar los últimos 4 dígitos de la tarjeta."), 400
            crudos = rex.parse_card_statement(destino, ultimos4)
            movimientos = [rex.normalizar_card(m, marca) for m in crudos]
        else:
            return jsonify(ok=False, error=f"Tipo de archivo desconocido: {tipo!r}"), 400
    except Exception as e:
        return jsonify(ok=False, error=f"No pude leer el archivo: {e}"), 400

    if not movimientos:
        return jsonify(ok=False, error="No encontré movimientos en ese archivo."), 400

    with db.conexion() as conn:
        db.crear_esquema(conn)
        stats = db.insertar_movimientos(conn, movimientos, origen=f"upload_{tipo}", usuario_id=viendo_id())

    return jsonify(ok=True, **stats)


def _leer_excel_generico(path) -> list[dict]:
    """Lee un .xlsx con las mismas columnas que la tabla movimientos
    (fecha, tipo, categoria, moneda, monto, descripcion, entidad), en la
    primera hoja del archivo."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h else h for h in row]
            continue
        if row[0] is None:
            continue
        d = dict(zip(headers, row))

        fecha = d.get("fecha")
        if isinstance(fecha, (datetime.datetime, datetime.date)):
            d["fecha"] = fecha.strftime("%Y-%m-%d")
        elif fecha is not None:
            d["fecha"] = str(fecha)

        monto = d.get("monto")
        d["monto"] = float(monto) if monto is not None else 0.0

        for campo in db.CAMPOS_ESPERADOS:
            d.setdefault(campo, "")

        rows.append(d)
    return rows
