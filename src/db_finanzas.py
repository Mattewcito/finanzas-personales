"""
db_finanzas.py
================
Módulo compartido de acceso a datos para el proyecto de finanzas personales.

FASE 0 del rediseño: mover la fuente de verdad de un Excel a una base de
datos SQLite real, local, sin dependencias externas (sqlite3 viene incluido
en Python — no hay que instalar nada).

Por ahora (transición), el Excel `data/finanzas_personales.xlsx` sigue
siendo lo que actualiza el bot de Gmail todos los días. Este módulo
SINCRONIZA desde ese Excel hacia `data/finanzas.db` cada vez que corre
`actualizar_dashboard.py`, así que la base de datos siempre queda al día
sin que nadie tenga que tocarla a mano. Cuando la lectura de correo pase a
ser una automatización local propia (próxima fase), la sincronización
dejará de ser "borrar y volver a cargar" y pasará a ser un merge con dedup
real por (fecha, monto).

Tablas:
  - movimientos: un registro por movimiento financiero, ya enriquecido con
    medio_pago y es_deuda (no hace falta recalcularlo en cada consulta).
  - historial_actualizaciones: espejo de la hoja del mismo nombre en el Excel.

Vista:
  - v_deuda_ledger: ledger cronológico de deuda de tarjeta de crédito con
    saldo corriente, calculado con una función de ventana SQL.
"""

import re
import sqlite3
import datetime
from collections import defaultdict
from contextlib import contextmanager
from pathlib import Path
from werkzeug.security import generate_password_hash, check_password_hash

import cifrado

SRC_DIR = Path(__file__).resolve().parent          # .../Finanzas personales/src
PROJECT_ROOT = SRC_DIR.parent                       # .../Finanzas personales
DATA_DIR = PROJECT_ROOT / "data"
XLSX_PATH = DATA_DIR / "finanzas_personales.xlsx"
DB_PATH = DATA_DIR / "finanzas.db"
SHEET_MOVIMIENTOS = "movimientos"
SHEET_HISTORIAL = "historial_actualizaciones"

CAMPOS_ESPERADOS = ("fecha", "tipo", "categoria", "moneda", "monto", "descripcion", "entidad")

ESQUEMA_SQL = """
CREATE TABLE IF NOT EXISTS movimientos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    fecha TEXT NOT NULL,
    tipo TEXT NOT NULL,
    categoria TEXT,
    moneda TEXT NOT NULL DEFAULT 'COP',
    monto REAL NOT NULL,
    descripcion TEXT,
    entidad TEXT,
    medio_pago TEXT NOT NULL DEFAULT 'debito',
    es_deuda INTEGER NOT NULL DEFAULT 0,
    origen TEXT NOT NULL DEFAULT 'gmail_bot_excel',
    referencia_bancaria TEXT,
    creado_en TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
);

CREATE INDEX IF NOT EXISTS idx_movimientos_fecha ON movimientos(fecha);
CREATE INDEX IF NOT EXISTS idx_movimientos_fecha_monto ON movimientos(fecha, monto);
CREATE INDEX IF NOT EXISTS idx_movimientos_origen ON movimientos(origen);

CREATE TABLE IF NOT EXISTS historial_actualizaciones (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    fecha_actualizacion TEXT NOT NULL,
    fecha_inicio_importada TEXT,
    fecha_fin_importada TEXT,
    movimientos_agregados INTEGER,
    origen TEXT DEFAULT 'gmail_bot_excel'
);

CREATE TABLE IF NOT EXISTS usuarios (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT NOT NULL,
    password_hash TEXT NOT NULL,
    rol TEXT NOT NULL DEFAULT 'usuario',   -- 'admin' | 'usuario'
    nombre_mostrado TEXT,
    creado_en TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
);
CREATE INDEX IF NOT EXISTS idx_usuarios_username ON usuarios(username);

-- Configuración de lectura de correo (Fase 1), UNA fila por usuario --
-- cada quien configura su propio correo dedicado desde "Mi perfil" en la
-- interfaz (routes/correo.py). app_password se guarda en texto plano a
-- propósito (no se puede hashear: leer_correo.py necesita el valor real
-- para autenticarse por IMAP) -- mismo nivel de confianza que ya tenía
-- data/credenciales_correo.json (archivo local, fuera de git, en una app
-- que nunca se expone a internet público). La interfaz nunca la vuelve a
-- mostrar una vez guardada.
CREATE TABLE IF NOT EXISTS correo_config (
    usuario_id INTEGER PRIMARY KEY REFERENCES usuarios(id),
    email TEXT NOT NULL,
    app_password TEXT NOT NULL,
    imap_host TEXT NOT NULL DEFAULT 'imap.gmail.com',
    imap_port INTEGER NOT NULL DEFAULT 993,
    cedula TEXT,          -- opcional: contraseña de los PDF de extracto adjuntos (Bancolombia los cifra con la cédula del titular)
    activo INTEGER NOT NULL DEFAULT 1,
    frecuencia_tipo TEXT NOT NULL DEFAULT 'intervalo',   -- 'intervalo' | 'diario'
    frecuencia_minutos INTEGER NOT NULL DEFAULT 30,       -- usado si frecuencia_tipo='intervalo'
    frecuencia_hora TEXT,                                 -- 'HH:MM', usado si frecuencia_tipo='diario'
    ultima_corrida TEXT,
    ultima_corrida_ok INTEGER,
    ultimo_error TEXT,
    actualizado_en TEXT NOT NULL DEFAULT (datetime('now', 'localtime'))
);
"""

# username no tiene restricción UNIQUE a nivel de base de datos, para
# soportar configuraciones de cuentas fuera del caso estándar.

VISTA_DEUDA_SQL = """
CREATE VIEW IF NOT EXISTS v_deuda_ledger AS
SELECT
    id,
    usuario_id,
    fecha,
    medio_pago AS tipo_movimiento,
    monto,
    descripcion,
    entidad,
    SUM(CASE WHEN medio_pago = 'pago_tarjeta_credito' THEN -monto ELSE monto END)
        OVER (PARTITION BY usuario_id ORDER BY fecha, id ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) AS saldo_acumulado
FROM movimientos
WHERE medio_pago IN ('credito', 'avance_credito', 'pago_tarjeta_credito')
  AND moneda = 'COP'
ORDER BY fecha, id;
"""


def conectar() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


@contextmanager
def conexion():
    """Context manager sobre conectar(): garantiza conn.close() incluso
    si algo lanza una excepción en el medio. Reemplaza el patrón repetido
    `conn = db.conectar(); try: ...; finally: conn.close()` que se
    duplicaba en cada ruta de app.py -- mismo comportamiento, menos
    código repetido. Uso: `with db.conexion() as conn: ...`."""
    conn = conectar()
    try:
        yield conn
    finally:
        conn.close()


def _agregar_columna_si_falta(conn: sqlite3.Connection, tabla: str, columna: str, tipo_sql: str) -> None:
    """ALTER TABLE ... ADD COLUMN, tolerante a la carrera entre procesos:
    gunicorn arranca esta app con varios workers (ver Dockerfile), cada
    uno importa app.py por separado y cada uno corre crear_esquema() al
    boot -- si dos lo hacen casi al mismo tiempo, el chequeo previo de
    PRAGMA table_info() puede pasar en los dos ANTES de que cualquiera
    haya hecho el ALTER, y el segundo revienta con "duplicate column
    name" (esto pasó de verdad al desplegar referencia_bancaria).
    SQLite no tiene 'ADD COLUMN IF NOT EXISTS', así que se ataja acá:
    chequeo previo (evita el ALTER en el caso común) + tolerar el error
    puntual de "ya existe" si igual se cuela la carrera."""
    columnas = [r["name"] for r in conn.execute(f"PRAGMA table_info({tabla})")]
    if columna in columnas:
        return
    try:
        conn.execute(f"ALTER TABLE {tabla} ADD COLUMN {columna} {tipo_sql}")
        conn.commit()
    except sqlite3.OperationalError as e:
        if "duplicate column name" not in str(e):
            raise  # cualquier otro error sí debe reventar, no ocultarlo


def _migrar_columna_usuario_id(conn: sqlite3.Connection) -> None:
    """Cada movimiento pasa a pertenecer a un usuario (multiusuario,
    2026-09-05). Nula por defecto en filas viejas; se asigna al admin la
    primera vez que corre esta migración (ver
    asignar_movimientos_sin_dueno_a_admin)."""
    _agregar_columna_si_falta(conn, "movimientos", "usuario_id", "INTEGER")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_movimientos_usuario ON movimientos(usuario_id)")
    conn.commit()


def _migrar_columna_referencia_bancaria(conn: sqlite3.Connection) -> None:
    """Guarda la descripción "oficial" (del banco, por correo/PDF) cuando
    esa misma transacción ya existía como registro manual -- ver
    insertar_movimientos(). Nula en todo lo insertado antes de esta
    migración (2026-09-06) y en cualquier movimiento que nunca se haya
    conciliado contra una fuente automática."""
    _agregar_columna_si_falta(conn, "movimientos", "referencia_bancaria", "TEXT")


def _migrar_columna_cedula_correo_config(conn: sqlite3.Connection) -> None:
    """correo_config ya existía (2026-09-06) sin esta columna en cualquier
    BD real donde ya se hubiera guardado alguna configuración -- CREATE
    TABLE IF NOT EXISTS no la agrega sola a una tabla que ya existe."""
    _agregar_columna_si_falta(conn, "correo_config", "cedula", "TEXT")


def _migrar_cifrado_correo_config(conn: sqlite3.Connection) -> None:
    """Cifra en el lugar cualquier fila de correo_config que haya
    quedado en texto plano de ANTES de que existiera cifrado.py
    (2026-09-06) -- email/app_password/cedula pasan a estar cifrados
    tanto para filas nuevas (ver guardar_correo_config) como viejas.
    cifrado.esta_cifrado() hace que sea seguro correr esto una y otra
    vez sin volver a cifrar un valor ya cifrado (lo dejaría ilegible)."""
    filas = conn.execute("SELECT usuario_id, app_password, cedula, email FROM correo_config").fetchall()
    for fila in filas:
        cambios = {}
        for campo in ("app_password", "cedula", "email"):
            valor = fila[campo]
            if valor and not cifrado.esta_cifrado(valor):
                cambios[campo] = cifrado.cifrar(valor)
        if cambios:
            set_sql = ", ".join(f"{c} = ?" for c in cambios)
            conn.execute(f"UPDATE correo_config SET {set_sql} WHERE usuario_id = ?",
                         (*cambios.values(), fila["usuario_id"]))
    if filas:
        conn.commit()


def crear_esquema(conn: sqlite3.Connection) -> None:
    conn.executescript(ESQUEMA_SQL)
    _migrar_columna_usuario_id(conn)
    _migrar_columna_referencia_bancaria(conn)
    _migrar_columna_cedula_correo_config(conn)
    _migrar_cifrado_correo_config(conn)
    # DROP + recrear la vista: si ya existía de antes de agregar usuario_id
    # a su SELECT, "CREATE VIEW IF NOT EXISTS" no la actualiza sola.
    conn.execute("DROP VIEW IF EXISTS v_deuda_ledger")
    conn.executescript(VISTA_DEUDA_SQL)
    conn.commit()


# ----------------------------- Clasificación (misma lógica ya validada) -----------------------------

def clasificar_medio_pago(descripcion: str) -> str:
    """Infiere el medio de pago a partir de patrones de texto típicos de
    las notificaciones bancarias de Bancolombia/Nequi/Nu.

    Valores posibles:
      - 'avance_credito'       → adelanto de efectivo de una tarjeta de crédito hacia una cuenta
      - 'credito'              → compra cargada a tarjeta de crédito (deuda, no caja)
      - 'pago_tarjeta_credito' → pago que abona/salda una tarjeta de crédito (caja real)
      - 'debito'               → todo lo demás: débito, efectivo, transferencias, QR, nómina, etc.
    """
    d = (descripcion or "").lower()
    if "avance" in d and ("t.cred" in d or "tcred" in d or "tarjeta de cr" in d):
        return "avance_credito"
    if "t.cred" in d or "tcred" in d:
        return "credito"
    if "pago tarjeta" in d or "pago tdc" in d or "pago t.cred" in d or "pago a tarjeta" in d:
        return "pago_tarjeta_credito"
    return "debito"


def enriquecer_movimiento(row: dict) -> dict:
    """Agrega 'medio_pago' y 'es_deuda', y reclasifica los casos especiales
    acordados: avance de crédito pasa a 'ingreso' marcado como deuda; compra
    a crédito sigue siendo 'gasto' pero marcada como deuda (no caja real
    todavía); pago de tarjeta se recategoriza y es_deuda=False (sí es caja
    real, liquida deuda)."""
    row = dict(row)
    medio = clasificar_medio_pago(str(row.get("descripcion", "")))
    row["medio_pago"] = medio

    if medio == "avance_credito":
        row["tipo"] = "ingreso"
        row["categoria"] = "avance_credito"
        row["es_deuda"] = True
    elif medio == "credito":
        row["es_deuda"] = (row.get("tipo") == "gasto")
    elif medio == "pago_tarjeta_credito":
        row["categoria"] = "pago_tarjeta_credito"
        row["es_deuda"] = False
    else:
        row["es_deuda"] = False

    return row


# ----------------------------- Lectura del Excel -----------------------------

def leer_movimientos_excel() -> list[dict]:
    """Lee la hoja 'movimientos' del Excel (solo lectura, nunca escribe)."""
    import openpyxl

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET_MOVIMIENTOS not in wb.sheetnames:
        raise ValueError(f"La hoja '{SHEET_MOVIMIENTOS}' no existe en {XLSX_PATH.name}.")
    ws = wb[SHEET_MOVIMIENTOS]

    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h else h for h in row]
            continue
        if row[0] is None:
            continue
        d = dict(zip(headers, row))

        fecha = d.get("fecha")
        if isinstance(fecha, (datetime.datetime, datetime.date)):
            d["fecha"] = fecha.strftime("%Y-%m-%d")
        elif fecha is not None:
            d["fecha"] = str(fecha)

        monto = d.get("monto")
        d["monto"] = float(monto) if monto is not None else 0.0

        for campo in CAMPOS_ESPERADOS:
            d.setdefault(campo, "")

        rows.append(d)

    if not rows:
        raise ValueError("No se encontraron movimientos en la hoja — el Excel podría estar vacío o mal formado.")
    return rows


def leer_historial_excel() -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET_HISTORIAL not in wb.sheetnames:
        return []
    ws = wb[SHEET_HISTORIAL]

    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h else h for h in row]
            continue
        if row[0] is None:
            continue
        d = dict(zip(headers, row))
        for k in ("fecha_actualizacion", "fecha_inicio_importada", "fecha_fin_importada"):
            v = d.get(k)
            if isinstance(v, (datetime.datetime, datetime.date)):
                d[k] = v.strftime("%Y-%m-%d %H:%M:%S") if isinstance(v, datetime.datetime) else v.strftime("%Y-%m-%d")
        rows.append(d)
    return rows


# ----------------------------- Sincronización Excel -> SQLite -----------------------------

def sincronizar_desde_excel(conn: sqlite3.Connection) -> dict:
    """Reemplaza el contenido de origen='gmail_bot_excel' con lo que hay
    ahora mismo en el Excel. Mientras el Excel sea la única fuente de
    escritura (fase actual), esto es seguro y simple: el Excel manda.
    Cuando existan otras fuentes de ingesta, dejará de borrar-y-recargar y
    pasará a hacer un merge con dedup por (fecha, monto)."""
    # El Excel es de una sola persona (el dueño original de la app) -- todo
    # lo que sincroniza acá se le asigna a la cuenta 'admin'.
    admin = conn.execute("SELECT id FROM usuarios WHERE rol = 'admin' LIMIT 1").fetchone()
    admin_id = admin["id"] if admin else None

    movimientos = [enriquecer_movimiento(m) for m in leer_movimientos_excel()]
    for m in movimientos:
        m["usuario_id"] = admin_id
    historial = leer_historial_excel()

    cur = conn.cursor()
    cur.execute("DELETE FROM movimientos WHERE origen = 'gmail_bot_excel'")
    cur.executemany(
        """INSERT INTO movimientos (fecha, tipo, categoria, moneda, monto, descripcion, entidad, medio_pago, es_deuda, origen, usuario_id)
           VALUES (:fecha, :tipo, :categoria, :moneda, :monto, :descripcion, :entidad, :medio_pago, :es_deuda, 'gmail_bot_excel', :usuario_id)""",
        movimientos,
    )

    cur.execute("DELETE FROM historial_actualizaciones WHERE origen = 'gmail_bot_excel'")
    for h in historial:
        cur.execute(
            """INSERT INTO historial_actualizaciones
               (fecha_actualizacion, fecha_inicio_importada, fecha_fin_importada, movimientos_agregados, origen)
               VALUES (?, ?, ?, ?, 'gmail_bot_excel')""",
            (
                h.get("fecha_actualizacion"),
                h.get("fecha_inicio_importada"),
                h.get("fecha_fin_importada"),
                h.get("movimientos_agregados"),
            ),
        )
    conn.commit()

    return {
        "movimientos": len(movimientos),
        "historial": len(historial),
        "fecha_min": min((m["fecha"] for m in movimientos if m["fecha"]), default=None),
        "fecha_max": max((m["fecha"] for m in movimientos if m["fecha"]), default=None),
        "en_deuda": sum(1 for m in movimientos if m["es_deuda"]),
    }


# ----------------------------- Consultas para el dashboard -----------------------------

def obtener_movimientos(conn: sqlite3.Connection, usuario_id: int | None = None) -> list[dict]:
    """usuario_id=None trae TODO (uso interno/admin explícito) -- las
    rutas de la app siempre deben pasar un usuario_id real."""
    # origen/creado_en/referencia_bancaria: para poder distinguir en el
    # dashboard qué se cargó a mano y qué llegó solo (correo/PDF/Excel),
    # y cuándo -- ayuda a decidir/confiar en cada movimiento, no solo a
    # verlo (pedido explícito 2026-09-06).
    sql = """SELECT id, fecha, tipo, categoria, moneda, monto, descripcion, entidad, medio_pago, es_deuda,
                    origen, referencia_bancaria, creado_en
             FROM movimientos"""
    params = ()
    if usuario_id is not None:
        sql += " WHERE usuario_id = ?"
        params = (usuario_id,)
    sql += " ORDER BY fecha, id"

    out = []
    for r in conn.execute(sql, params).fetchall():
        d = dict(r)
        d["es_deuda"] = bool(d["es_deuda"])
        out.append(d)
    return out


# ----------------------------- Usuarios / login -----------------------------

def crear_usuario(conn: sqlite3.Connection, username: str, password: str, rol: str, nombre_mostrado: str) -> int:
    """La contraseña se guarda SIEMPRE hasheada (nunca en texto plano),
    con el algoritmo por defecto de Werkzeug (scrypt)."""
    cur = conn.execute(
        "INSERT INTO usuarios (username, password_hash, rol, nombre_mostrado) VALUES (?, ?, ?, ?)",
        (username, generate_password_hash(password), rol, nombre_mostrado),
    )
    conn.commit()
    return cur.lastrowid


def actualizar_usuario(conn: sqlite3.Connection, usuario_id: int, username: str = None,
                        nombre_mostrado: str = None, password: str = None, rol: str = None) -> None:
    """Actualiza solo los campos que se pasen (None = no tocar ese campo).
    La contraseña, si se pasa, se hashea acá mismo -- nunca se guarda en
    texto plano, igual que en crear_usuario()."""
    campos, valores = [], []
    if username is not None:
        campos.append("username = ?"); valores.append(username)
    if nombre_mostrado is not None:
        campos.append("nombre_mostrado = ?"); valores.append(nombre_mostrado)
    if rol is not None:
        campos.append("rol = ?"); valores.append(rol)
    if password:
        campos.append("password_hash = ?"); valores.append(generate_password_hash(password))
    if not campos:
        return
    valores.append(usuario_id)
    conn.execute(f"UPDATE usuarios SET {', '.join(campos)} WHERE id = ?", valores)
    conn.commit()


def verificar_login(conn: sqlite3.Connection, username: str, password: str) -> dict | None:
    """Verifica las credenciales contra las cuentas registradas con ese
    nombre de usuario y devuelve la cuenta correspondiente si coinciden."""
    candidatos = conn.execute("SELECT * FROM usuarios WHERE username = ?", (username,)).fetchall()
    for c in candidatos:
        if check_password_hash(c["password_hash"], password):
            return dict(c)
    return None


def listar_usuarios(conn: sqlite3.Connection) -> list[dict]:
    return [dict(r) for r in conn.execute("SELECT id, username, rol, nombre_mostrado FROM usuarios ORDER BY id")]


# ----------------------------- Configuración de lectura de correo -----------------------------
# email/app_password/cedula se guardan CIFRADOS en la columna (ver
# cifrado.py) -- esta es la ÚNICA capa que cifra/descifra; el resto del
# código (routes/correo.py, leer_correo.py) siempre trabaja con texto
# plano en memoria, como si el cifrado no existiera.

def _descifrar_fila_correo(fila: dict) -> dict:
    """Si algún campo no se puede descifrar (clave distinta, dato
    corrupto), queda en None en vez de tumbar toda la cuenta -- el login
    IMAP/PDF va a fallar igual con un error claro más adelante, pero no
    rompe el procesamiento de las DEMÁS cuentas en listar_correo_configs_activos()."""
    fila = dict(fila)
    for campo in ("email", "app_password", "cedula"):
        valor = fila.get(campo)
        if valor:
            try:
                fila[campo] = cifrado.descifrar(valor)
            except ValueError:
                fila[campo] = None
    return fila


def obtener_correo_config(conn: sqlite3.Connection, usuario_id: int) -> dict | None:
    r = conn.execute("SELECT * FROM correo_config WHERE usuario_id = ?", (usuario_id,)).fetchone()
    return _descifrar_fila_correo(r) if r else None


def listar_correo_configs_activos(conn: sqlite3.Connection) -> list[dict]:
    """Todas las cuentas con la automatización encendida -- lo que
    leer_correo.py recorre en cada corrida de la tarea programada."""
    return [_descifrar_fila_correo(r) for r in conn.execute("SELECT * FROM correo_config WHERE activo = 1")]


def guardar_correo_config(
    conn: sqlite3.Connection,
    usuario_id: int,
    email: str,
    app_password: str | None = None,
    imap_host: str = "imap.gmail.com",
    imap_port: int = 993,
    cedula: str | None = None,
    frecuencia_tipo: str = "intervalo",
    frecuencia_minutos: int = 30,
    frecuencia_hora: str | None = None,
    activo: bool = True,
) -> None:
    """Crea o actualiza la configuración de correo de un usuario (una fila
    por usuario_id). `app_password=None` (o vacío) significa "no cambiar
    la que ya había guardada" -- así el formulario de edición no obliga a
    reescribirla cada vez que se toca cualquier otro campo (ej. la
    frecuencia). Es obligatoria la primera vez que se guarda esta cuenta.
    `cedula` es opcional (solo hace falta si se quiere que también se
    abran PDFs de extracto adjuntos, cifrados con ese número) y sigue el
    mismo criterio: vacío = no tocar la que ya había."""
    existente = obtener_correo_config(conn, usuario_id)  # ya viene descifrado (ver _descifrar_fila_correo)
    if not app_password:
        if not existente:
            raise ValueError("Falta la contraseña de aplicación (obligatoria la primera vez que se configura).")
        app_password = existente["app_password"]
    if not cedula and existente:
        cedula = existente["cedula"]

    conn.execute(
        """
        INSERT INTO correo_config
            (usuario_id, email, app_password, imap_host, imap_port, cedula, activo,
             frecuencia_tipo, frecuencia_minutos, frecuencia_hora, actualizado_en)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now', 'localtime'))
        ON CONFLICT(usuario_id) DO UPDATE SET
            email = excluded.email,
            app_password = excluded.app_password,
            imap_host = excluded.imap_host,
            imap_port = excluded.imap_port,
            cedula = excluded.cedula,
            activo = excluded.activo,
            frecuencia_tipo = excluded.frecuencia_tipo,
            frecuencia_minutos = excluded.frecuencia_minutos,
            frecuencia_hora = excluded.frecuencia_hora,
            actualizado_en = excluded.actualizado_en
        """,
        (usuario_id, cifrado.cifrar(email), cifrado.cifrar(app_password), imap_host, imap_port,
         cifrado.cifrar(cedula), int(bool(activo)), frecuencia_tipo, frecuencia_minutos, frecuencia_hora),
    )
    conn.commit()


def actualizar_estado_correo(conn: sqlite3.Connection, usuario_id: int, ok: bool, error: str | None = None) -> None:
    """Deja constancia del resultado de la última corrida -- lo que se
    muestra en la interfaz ("última sincronización: hace 12 min, OK")."""
    conn.execute(
        "UPDATE correo_config SET ultima_corrida = datetime('now', 'localtime'), "
        "ultima_corrida_ok = ?, ultimo_error = ? WHERE usuario_id = ?",
        (int(bool(ok)), error, usuario_id),
    )
    conn.commit()


def eliminar_correo_config(conn: sqlite3.Connection, usuario_id: int) -> None:
    conn.execute("DELETE FROM correo_config WHERE usuario_id = ?", (usuario_id,))
    conn.commit()


def obtener_usuario(conn: sqlite3.Connection, usuario_id: int) -> dict | None:
    r = conn.execute("SELECT id, username, rol, nombre_mostrado FROM usuarios WHERE id = ?", (usuario_id,)).fetchone()
    return dict(r) if r else None


def asignar_movimientos_sin_dueno_a_admin(conn: sqlite3.Connection) -> int:
    """Los movimientos que ya existían antes del sistema de usuarios
    (usuario_id NULL) son todos del dueño original de la app -- se le
    asignan a la cuenta 'admin' la primera vez que corre esto."""
    admin = conn.execute("SELECT id FROM usuarios WHERE rol = 'admin' LIMIT 1").fetchone()
    if not admin:
        return 0
    cur = conn.execute("UPDATE movimientos SET usuario_id = ? WHERE usuario_id IS NULL", (admin["id"],))
    conn.commit()
    return cur.rowcount


def obtener_categorias(conn: sqlite3.Connection, usuario_id: int | None = None) -> list[str]:
    """Categorías distintas ya usadas, para autocompletar el formulario de
    registro manual (evita que cada quien escriba la misma categoría con
    variantes distintas). Filtradas por usuario: cada quien autocompleta
    con SU propio historial."""
    sql = "SELECT DISTINCT categoria FROM movimientos WHERE categoria IS NOT NULL AND categoria != ''"
    params = ()
    if usuario_id is not None:
        sql += " AND usuario_id = ?"
        params = (usuario_id,)
    sql += " ORDER BY categoria"
    return [r["categoria"] for r in conn.execute(sql, params).fetchall()]


def obtener_entidades(conn: sqlite3.Connection, usuario_id: int | None = None) -> list[str]:
    sql = "SELECT DISTINCT entidad FROM movimientos WHERE entidad IS NOT NULL AND entidad != ''"
    params = ()
    if usuario_id is not None:
        sql += " AND usuario_id = ?"
        params = (usuario_id,)
    sql += " ORDER BY entidad"
    return [r["entidad"] for r in conn.execute(sql, params).fetchall()]


def obtener_ledger_deuda(conn: sqlite3.Connection, usuario_id: int | None = None) -> list[dict]:
    sql = "SELECT fecha, tipo_movimiento, monto, descripcion, entidad, saldo_acumulado FROM v_deuda_ledger"
    params = ()
    if usuario_id is not None:
        sql += " WHERE usuario_id = ?"
        params = (usuario_id,)
    return [dict(r) for r in conn.execute(sql, params).fetchall()]


# ----------------------------- Inserción con dedup (usada por cualquier fuente de ingesta) -----------------------------

def _palabras(*textos: str) -> set[str]:
    return set(re.findall(r"\w+", " ".join(t or "" for t in textos).lower()))


def _mejor_coincidencia(m: dict, candidatos: list[dict]) -> dict | None:
    """De entre los movimientos ya existentes que calzan en (moneda, monto,
    tipo), elige el más probable de ser LA MISMA transacción real que `m`:
    fecha exacta gana sobre ±1 día (el banco y el registro manual pueden
    quedar a caballo de la medianoche), y entre empates, el que más
    palabras comparte en descripción/entidad -- pero esto último es solo
    un DESEMPATE, nunca un requisito: un registro manual ("almuerzo") y
    uno de correo ("Transferiste $30.000 de tu cuenta...") legítimamente
    no comparten ninguna palabra siendo la misma transacción."""
    fecha_m = datetime.date.fromisoformat(m["fecha"])
    en_ventana = [
        c for c in candidatos
        if abs((datetime.date.fromisoformat(c["fecha"]) - fecha_m).days) <= 1
    ]
    if not en_ventana:
        return None
    palabras_m = _palabras(m.get("descripcion", ""), m.get("entidad", ""))

    def _orden(c: dict):
        fecha_exacta = 0 if c["fecha"] == m["fecha"] else 1
        solapadas = len(palabras_m & _palabras(c.get("descripcion", ""), c.get("entidad", "")))
        return (fecha_exacta, -solapadas, c["id"])

    return min(en_ventana, key=_orden)


def insertar_movimientos(conn: sqlite3.Connection, movimientos: list[dict], origen: str, usuario_id: int) -> dict:
    """Inserta los movimientos de `movimientos` que no correspondan a una
    transacción YA registrada para ese usuario -- cada usuario tiene sus
    propias finanzas separadas, así que lo de otro usuario nunca cuenta
    como duplicado. Cada movimiento debe traer al menos
    fecha/tipo/categoria/moneda/monto/descripcion/entidad; se enriquece
    acá (medio_pago/es_deuda) antes de insertar.

    CONCILIACIÓN (2026-09-06): el match ya no es "existe alguna fila con
    esa fecha+monto" (eso confundía dos transacciones reales distintas
    que coincidieran en fecha+monto -- ej. dos transferencias de $30.000
    el mismo día -- tratando la segunda como si fuera repetida). Ahora
    cada fila existente solo puede "absorber" UN incoming (multiset, no
    set): se consume al usarse, así que una segunda coincidencia real
    ese mismo día sí se inserta como nueva. Además, cuando lo que
    absorbe la coincidencia es un registro CARGADO A MANO (origen
    'app_manual') y lo que llega es de una fuente automática (correo,
    PDF, Excel), se guarda la descripción "oficial" del banco en
    referencia_bancaria SIN tocar la descripción/categoría que el
    usuario ya había escrito (ej. "almuerzo" se queda como está).

    El "consumo" tiene que sobrevivir entre llamadas distintas (cada
    corrida de leer_correo.py es una llamada separada) -- por eso una
    fila manual que YA absorbió una coincidencia (referencia_bancaria ya
    no es NULL) queda EXCLUIDA de volver a ser candidata en el futuro:
    si no, una segunda transacción real de $30.000 el mismo día,
    detectada en una corrida posterior, volvería a "encontrar" la misma
    fila de "almuerzo" y se perdería en vez de insertarse. Una fila
    automática ya existente sí puede seguir absorbiendo coincidencias
    indefinidamente entre corridas -- eso es re-detectar el mismo correo
    dos veces (ventanas que se solapan), que sí debe seguir marcándose
    como duplicado siempre."""
    cur = conn.cursor()
    existentes = [
        dict(r) for r in cur.execute(
            "SELECT id, fecha, moneda, monto, tipo, descripcion, entidad, origen, referencia_bancaria "
            "FROM movimientos WHERE usuario_id = ? "
            "AND NOT (origen = 'app_manual' AND referencia_bancaria IS NOT NULL)",
            (usuario_id,),
        )
    ]
    disponibles: dict[tuple, list[dict]] = defaultdict(list)
    for e in existentes:
        disponibles[(e["moneda"], round(e["monto"]), e["tipo"])].append(e)

    # Se distinguen dos tipos de duplicado -- son casos distintos y el
    # mensaje al usuario debe aclarar cuál es cuál:
    #   - duplicados_bd: esa transacción ya estaba guardada de antes.
    #   - duplicados_lote: dos filas DEL MISMO archivo/lote son iguales
    #     entre sí (no existían antes, pero no tiene sentido guardar la
    #     misma dos veces en la misma carga).
    vistos_en_lote = set()
    nuevos, duplicados_bd, duplicados_lote = [], 0, 0
    for m_crudo in movimientos:
        # Enriquecer ANTES de armar la clave de match: enriquecer_movimiento
        # puede reclasificar 'tipo' (ej. avance de tarjeta: 'gasto' ->
        # 'ingreso'), y las filas ya existentes en `disponibles` tienen el
        # tipo YA reclasificado -- si acá se comparara con el tipo crudo,
        # un avance nunca encontraría su propia fila ya guardada y se
        # insertaría de nuevo en cada corrida.
        m = enriquecer_movimiento(m_crudo)

        clave_lote = (m["fecha"], m["moneda"], round(m["monto"]), m["tipo"])
        if clave_lote in vistos_en_lote:
            duplicados_lote += 1
            continue

        candidatos = disponibles.get((m["moneda"], round(m["monto"]), m["tipo"]), [])
        match = _mejor_coincidencia(m, candidatos) if candidatos else None
        if match:
            candidatos.remove(match)  # consumida -- una segunda coincidencia real no la vuelve a encontrar
            duplicados_bd += 1
            if match["origen"] == "app_manual" and origen != "app_manual" and not match.get("referencia_bancaria"):
                referencia = (m.get("descripcion") or "").strip()
                if referencia:
                    cur.execute("UPDATE movimientos SET referencia_bancaria = ? WHERE id = ?", (referencia, match["id"]))
            continue

        vistos_en_lote.add(clave_lote)
        nuevos.append(m)

    for e in nuevos:
        e["origen"] = origen
        e["usuario_id"] = usuario_id
    cur.executemany(
        """INSERT INTO movimientos (fecha, tipo, categoria, moneda, monto, descripcion, entidad, medio_pago, es_deuda, origen, usuario_id)
           VALUES (:fecha, :tipo, :categoria, :moneda, :monto, :descripcion, :entidad, :medio_pago, :es_deuda, :origen, :usuario_id)""",
        nuevos,
    )
    conn.commit()
    return {
        "nuevos": len(nuevos),
        "duplicados": duplicados_bd + duplicados_lote,
        "duplicados_bd": duplicados_bd,
        "duplicados_lote": duplicados_lote,
    }
